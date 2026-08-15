import io
from datetime import datetime
from flask import Blueprint, request, send_file
from flask_jwt_extended import jwt_required
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from app import db
from app.models.member import Member

reports_bp = Blueprint("reports", __name__)


def _get_members(args):
    status = args.get("status", "").strip()
    search = args.get("search", "").strip()
    query = Member.query
    if status:
        query = query.filter(Member.status == status)
    if search:
        like = f"%{search}%"
        query = query.filter(
            db.or_(
                Member.first_name.ilike(like),
                Member.last_name.ilike(like),
                Member.member_number.ilike(like),
            )
        )
    return query.order_by(Member.last_name, Member.first_name).all()


@reports_bp.get("/pdf")
@jwt_required()
def export_pdf():
    members = _get_members(request.args)
    buffer = io.BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=15 * mm,
        leftMargin=15 * mm,
        topMargin=20 * mm,
        bottomMargin=15 * mm,
    )

    styles = getSampleStyleSheet()
    elements = []

    title = Paragraph(
        f"<b>ZBC Church — Member Report</b><br/><font size='9'>Generated: {datetime.now().strftime('%d %B %Y %H:%M')}</font>",
        styles["Title"],
    )
    elements.append(title)
    elements.append(Spacer(1, 8 * mm))

    headers = ["#", "Member No.", "Full Name", "Gender", "Phone", "Email", "Status", "Join Date"]
    table_data = [headers]

    for i, m in enumerate(members, start=1):
        table_data.append([
            str(i),
            m.member_number,
            f"{m.first_name} {m.last_name}",
            (m.gender or "").capitalize(),
            m.phone or "",
            m.email or "",
            (m.status or "").capitalize(),
            m.join_date.strftime("%d/%m/%Y") if m.join_date else "",
        ])

    col_widths = [10 * mm, 30 * mm, 50 * mm, 20 * mm, 35 * mm, 55 * mm, 22 * mm, 28 * mm]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e40af")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#eff6ff")]),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))

    elements.append(table)
    doc.build(elements)
    buffer.seek(0)

    filename = f"members_{datetime.now().strftime('%Y%m%d')}.pdf"
    return send_file(buffer, mimetype="application/pdf", as_attachment=True, download_name=filename)


@reports_bp.get("/excel")
@jwt_required()
def export_excel():
    members = _get_members(request.args)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Members"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1E40AF")
    header_align = Alignment(horizontal="center", vertical="center")

    headers = [
        "No.", "Member Number", "First Name", "Last Name", "Date of Birth",
        "Gender", "Phone", "Email", "Address", "Status", "Join Date", "Departments",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    alt_fill = PatternFill("solid", fgColor="EFF6FF")

    for i, m in enumerate(members, start=1):
        dept_names = ", ".join(d.name for d in m.departments)
        row = [
            i,
            m.member_number,
            m.first_name,
            m.last_name,
            m.dob.strftime("%d/%m/%Y") if m.dob else "",
            (m.gender or "").capitalize(),
            m.phone or "",
            m.email or "",
            m.address or "",
            (m.status or "").capitalize(),
            m.join_date.strftime("%d/%m/%Y") if m.join_date else "",
            dept_names,
        ]
        ws.append(row)
        if i % 2 == 0:
            for cell in ws[i + 1]:
                cell.fill = alt_fill

    col_widths = [6, 18, 16, 16, 14, 10, 16, 28, 30, 12, 14, 30]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"members_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(
        buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@reports_bp.get("/stats")
@jwt_required()
def get_stats():
    from sqlalchemy import func
    from datetime import date

    total = Member.query.count()
    active = Member.query.filter_by(status="active").count()
    inactive = Member.query.filter_by(status="inactive").count()
    visitors = Member.query.filter_by(status="visitor").count()

    this_month = Member.query.filter(
        func.extract("month", Member.created_at) == date.today().month,
        func.extract("year", Member.created_at) == date.today().year,
    ).count()

    return {
        "total": total,
        "active": active,
        "inactive": inactive,
        "visitors": visitors,
        "new_this_month": this_month,
    }, 200
